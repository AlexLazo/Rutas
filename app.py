from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os
import pandas as pd
from datetime import datetime, time
import json
from functools import wraps
from pathlib import Path

# Importar configuración para Railway
try:
    from railway_config import get_port, get_debug_mode, print_env_info
    RAILWAY_CONFIG_AVAILABLE = True
except ImportError:
    RAILWAY_CONFIG_AVAILABLE = False

app = Flask(__name__)
app.secret_key = 'clave-secreta-rutas-2024'  # Cambiar en producción

# Headers de seguridad para parecer sitio web corporativo normal
@app.after_request
def add_security_headers(response):
    """Agregar headers que hagan parecer el sitio como corporativo normal"""
    # ELIMINAR completamente headers que delatan hosting externo
    headers_to_remove = [
        'Server', 'X-Powered-By', 'x-railway-edge', 'x-railway-request-id', 
        'via', 'x-vercel-cache', 'x-vercel-id', 'cf-ray', 'cf-cache-status',
        'x-served-by', 'x-cache', 'x-timer', 'x-fastly-request-id'
    ]
    
    for header in headers_to_remove:
        response.headers.pop(header, None)
        response.headers.pop(header.upper(), None)
        response.headers.pop(header.lower(), None)
    
    # Headers que simulan un servidor corporativo Microsoft/IIS INTERNO
    response.headers['Server'] = 'Microsoft-IIS/10.0'
    response.headers['X-Powered-By'] = 'ASP.NET'
    response.headers['X-AspNet-Version'] = '4.0.30319'
    response.headers['X-AspNetMvc-Version'] = '5.2'
    
    # Headers de seguridad estándar corporativo
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    
    # Headers que hacen parecer un sistema empresarial INTERNO
    response.headers['X-Enterprise-System'] = 'Sistema-Gestion-Rutas'
    response.headers['X-Company-Domain'] = 'distribuidora-corporativa.com'
    response.headers['X-Corporate-Network'] = 'INTERNAL'
    response.headers['X-System-Environment'] = 'PRODUCTION-CORPORATE'
    response.headers['X-Authentication-Provider'] = 'ActiveDirectory-Enterprise'
    response.headers['X-Corporate-Gateway'] = 'CORP-DMZ-01'
    response.headers['X-Internal-Service'] = 'Route-Management-System'
    
    # Headers adicionales para Render.com
    response.headers['X-Hosting-Platform'] = 'Render-Enterprise'
    response.headers['X-Deployment-Environment'] = 'Corporate-Cloud'
    
    # Headers específicos para parecer legítimo
    response.headers['X-Division'] = 'Logistica-Distribucion'
    response.headers['X-Region'] = 'CENTROAMERICA'
    response.headers['X-System-ID'] = 'RMS-CORP-2024'
    response.headers['X-Corporate-Auth'] = 'Integrated-Windows-Auth'
    
    # Headers de cache corporativo INTERNO
    response.headers['Cache-Control'] = 'private, no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    return response

# Configurar Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicie sesión para acceder a esta página.'
login_manager.login_message_category = 'info'

# Clase User para Flask-Login
class User(UserMixin):
    def __init__(self, id, username, email, role, active=True):
        self.id = id
        self.username = username
        self.email = email
        self.role = role
        self._is_active = active
    
    @property
    def is_active(self):
        return self._is_active
    
    def get_id(self):
        return str(self.id)
    
    def is_admin(self):
        return self.role in ['admin', 'super_admin']
    
    def is_supervisor(self):
        return self.role in ['supervisor', 'admin', 'super_admin']

@login_manager.user_loader
def load_user(user_id):
    """Cargar usuario desde la base de datos"""
    conn = get_db_connection()
    user_data = conn.execute(
        'SELECT * FROM users WHERE id = ? AND is_active = 1', 
        (user_id,)
    ).fetchone()
    conn.close()
    
    if user_data:
        return User(
            id=user_data['id'],
            username=user_data['username'],
            email=user_data['email'],
            role=user_data['role'],
            active=user_data['is_active']
        )
    return None

# Configuración de la base de datos
DATABASE = 'sistema_rutas.db'

def get_db_connection():
    """Obtener conexión a la base de datos"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_db():
    """Inicializar la base de datos"""
    print(f"🔄 Inicializando base de datos en: {DATABASE}")
    
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    # Configurar SQLite para mejor rendimiento
    cursor.execute("PRAGMA journal_mode=WAL")
    cursor.execute("PRAGMA synchronous=NORMAL")
    cursor.execute("PRAGMA foreign_keys=ON")
    cursor.execute("PRAGMA temp_store=MEMORY")
    
    # Tabla para almacenar datos de rutas (cargados desde Excel)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS rutas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ruta TEXT NOT NULL,
            codigo TEXT,
            placa TEXT,
            supervisor TEXT,
            contratista TEXT NOT NULL,
            tipo TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Tabla para reportes de rutas diarios
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reportes_rutas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contratista TEXT NOT NULL,
            ruta_id INTEGER NOT NULL,
            ruta_codigo TEXT NOT NULL,
            clientes_pendientes INTEGER NOT NULL DEFAULT 0,
            cajas_camion INTEGER NOT NULL DEFAULT 0,
            hora_aproximada_ingreso TIME NOT NULL,
            ubicacion_exacta TEXT,
            latitud REAL,
            longitud REAL,
            hora_exacta_envio DATETIME DEFAULT CURRENT_TIMESTAMP,
            comentarios TEXT,
            fecha_reporte DATE DEFAULT (date('now')),
            hora_reporte DATETIME DEFAULT CURRENT_TIMESTAMP,
            estado TEXT DEFAULT 'activo',
            reportado_por TEXT,
            FOREIGN KEY (ruta_id) REFERENCES rutas (id)
        )
    ''')
    
    # Tabla para usuarios del sistema
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            is_active INTEGER DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            last_login DATETIME,
            created_by INTEGER,
            FOREIGN KEY (created_by) REFERENCES users (id)
        )
    ''')
    
    # Tabla para log de actividades
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT NOT NULL,
            target_type TEXT,
            target_id INTEGER,
            details TEXT,
            ip_address TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    # Crear usuario administrador por defecto si no existe
    existing_admin = cursor.execute(
        'SELECT id FROM users WHERE username = ?', ('admin',)
    ).fetchone()
    
    if not existing_admin:
        admin_password_hash = generate_password_hash('admin123')
        cursor.execute('''
            INSERT INTO users (username, email, password_hash, role)
            VALUES (?, ?, ?, ?)
        ''', ('admin', 'admin@sistema-rutas.com', admin_password_hash, 'super_admin'))
        
        # Crear supervisor de ejemplo
        supervisor_password_hash = generate_password_hash('supervisor123')
        cursor.execute('''
            INSERT INTO users (username, email, password_hash, role, created_by)
            VALUES (?, ?, ?, ?, ?)
        ''', ('supervisor', 'supervisor@sistema-rutas.com', supervisor_password_hash, 'supervisor', 1))
        
        print("✅ Usuarios por defecto creados:")
        print("   Admin: admin / admin123")
        print("   Supervisor: supervisor / supervisor123")
    
    conn.commit()
    conn.close()

def load_rutas_from_excel():
    """Cargar rutas desde el archivo Excel"""
    try:
        # Leer el archivo Excel
        df = pd.read_excel('DB_Rutas.xlsx')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar si hay reportes que dependen de rutas
        reportes_count = cursor.execute('SELECT COUNT(*) FROM reportes_rutas').fetchone()[0]
        
        if reportes_count > 0:
            print(f"⚠️ No se pueden limpiar rutas: hay {reportes_count} reportes asociados")
            conn.close()
            return False
            
        # Limpiar tabla de rutas existente
        cursor.execute('DELETE FROM rutas')
        
        # Insertar datos desde Excel
        loaded_count = 0
        for _, row in df.iterrows():
            try:
                ruta = str(row['RUTA']) if pd.notna(row['RUTA']) else ''
                codigo = str(row['CODIGO']) if pd.notna(row['CODIGO']) else ''
                placa = str(row['PLACA']) if pd.notna(row['PLACA']) else ''
                supervisor = str(row['SUPERVISOR']) if pd.notna(row['SUPERVISOR']) else ''
                contratista = str(row['CONTRATISTA']) if pd.notna(row['CONTRATISTA']) else ''
                tipo = str(row['TIPO']) if pd.notna(row['TIPO']) else ''
                
                # Verificar que contratista no sea vacío (es NOT NULL)
                if not contratista:
                    print(f"⚠️ Saltando ruta sin contratista: {ruta}")
                    continue
                
                cursor.execute('''
                    INSERT INTO rutas (ruta, codigo, placa, supervisor, contratista, tipo)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (ruta, codigo, placa, supervisor, contratista, tipo))
                
                loaded_count += 1
            except Exception as row_error:
                print(f"⚠️ Error en fila {_}: {row_error}")
                continue
        
        conn.commit()
        conn.close()
        
        print(f"✅ {loaded_count} rutas cargadas desde Excel")
        return True
        
    except Exception as e:
        print(f"❌ Error cargando rutas desde Excel: {e}")
        return False

@app.route('/')
def index():
    """Página principal con el formulario para reportar rutas"""
    # Obtener contratistas únicos para el dropdown
    conn = get_db_connection()
    contratistas = conn.execute('''
        SELECT DISTINCT contratista 
        FROM rutas 
        WHERE contratista != '' 
        ORDER BY contratista
    ''').fetchall()
    conn.close()
    
    return render_template('index.html', contratistas=contratistas)

@app.route('/get_rutas/<contratista>')
def get_rutas(contratista):
    """API para obtener rutas de un contratista específico"""
    conn = get_db_connection()
    rutas = conn.execute('''
        SELECT id, ruta, codigo, supervisor, placa, tipo
        FROM rutas 
        WHERE contratista = ?
        ORDER BY ruta
    ''', (contratista,)).fetchall()
    conn.close()
    
    rutas_list = []
    for ruta in rutas:
        rutas_list.append({
            'id': ruta['id'],
            'ruta': ruta['ruta'],
            'codigo': ruta['codigo'],
            'supervisor': ruta['supervisor'],
            'placa': ruta['placa'],
            'tipo': ruta['tipo']
        })
    
    return jsonify(rutas_list)

@app.route('/submit_reporte', methods=['POST'])
def submit_reporte():
    """Procesar el envío del reporte de ruta"""
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        required_fields = ['contratista', 'ruta_id', 'clientes_pendientes', 
                          'cajas_camion', 'hora_aproximada_ingreso']
        
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Campo requerido: {field}'}), 400
        
        # Validar formato de hora
        try:
            hora_obj = datetime.strptime(data['hora_aproximada_ingreso'], '%H:%M').time()
        except ValueError:
            return jsonify({'success': False, 'error': 'Formato de hora inválido. Use HH:MM'}), 400
        
        # Obtener información de la ruta
        conn = get_db_connection()
        ruta_info = conn.execute(
            'SELECT ruta FROM rutas WHERE id = ?', 
            (data['ruta_id'],)
        ).fetchone()
        
        if not ruta_info:
            conn.close()
            return jsonify({'success': False, 'error': 'Ruta no encontrada'}), 404
        
        # Insertar reporte
        cursor = conn.cursor()
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        hora_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        cursor.execute('''
            INSERT INTO reportes_rutas 
            (contratista, ruta_id, ruta_codigo, clientes_pendientes, 
             cajas_camion, hora_aproximada_ingreso, ubicacion_exacta, 
             latitud, longitud, hora_exacta_envio, comentarios, reportado_por,
             fecha_reporte, hora_reporte)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['contratista'],
            data['ruta_id'],
            ruta_info['ruta'],
            int(data['clientes_pendientes']),
            int(data['cajas_camion']),
            data['hora_aproximada_ingreso'],
            data.get('ubicacion_exacta', ''),
            data.get('latitud'),
            data.get('longitud'),
            hora_actual,  # Hora exacta de envío
            data.get('comentarios', ''),
            data.get('reportado_por', 'Sistema'),
            fecha_actual,  # Fecha de reporte explícita
            hora_actual   # Hora de reporte explícita
        ))
        
        reporte_id = cursor.lastrowid
        conn.commit()
        
        # Verificar que el reporte se guardó correctamente
        verificacion = conn.execute(
            'SELECT id, fecha_reporte, hora_reporte FROM reportes_rutas WHERE id = ?', 
            (reporte_id,)
        ).fetchone()
        
        conn.close()
        
        print(f"✅ REPORTE GUARDADO - ID: {reporte_id}")
        print(f"   Contratista: {data['contratista']}")
        print(f"   Ruta: {ruta_info['ruta']}")
        print(f"   Fecha guardada: {verificacion['fecha_reporte'] if verificacion else 'ERROR'}")
        print(f"   Hora guardada: {verificacion['hora_reporte'] if verificacion else 'ERROR'}")
        
        return jsonify({
            'success': True, 
            'message': 'Reporte de ruta enviado exitosamente',
            'reporte_id': reporte_id,
            'fecha_guardada': verificacion['fecha_reporte'] if verificacion else None
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin')
@login_required
def admin():
    """Panel de administración para ver reportes de rutas"""
    # Obtener fecha actual o la fecha del filtro si se proporciona
    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    fecha_filtro = request.args.get('fecha', fecha_actual)
    contratista_filtro = request.args.get('contratista', '')
    
    # Parámetros de paginación
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 10))  # 25 reportes por página
    offset = (page - 1) * per_page
    
    conn = get_db_connection()
    
    # Mostrar todas las tablas para diagnóstico
    tables = conn.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()
    print("Tablas en la base de datos:", [t[0] for t in tables])
    
    # Construir query base para contar total
    count_query = '''
        SELECT COUNT(*) 
        FROM reportes_rutas r
        LEFT JOIN rutas ru ON r.ruta_id = ru.id
        WHERE 1=1
    '''
    count_params = []
    
    if fecha_filtro:
        count_query += ' AND date(r.fecha_reporte) = date(?)'
        count_params.append(fecha_filtro)
    
    if contratista_filtro:
        count_query += ' AND r.contratista = ?'
        count_params.append(contratista_filtro)
    
    # Obtener total de reportes con filtros
    total_reportes = conn.execute(count_query, count_params).fetchone()[0]
    print(f"Total de reportes con filtros: {total_reportes}")
    
    # Construir query con filtros y paginación
    query = '''
        SELECT r.*, ru.supervisor, ru.placa, ru.tipo
        FROM reportes_rutas r
        LEFT JOIN rutas ru ON r.ruta_id = ru.id
        WHERE 1=1
    '''
    params = []
    
    if fecha_filtro:
        query += ' AND date(r.fecha_reporte) = date(?)'
        params.append(fecha_filtro)
    
    if contratista_filtro:
        query += ' AND r.contratista = ?'
        params.append(contratista_filtro)
    
    query += ' ORDER BY r.hora_reporte DESC LIMIT ? OFFSET ?'
    params.extend([per_page, offset])
    
    print(f"DEBUG: Ejecutando consulta con fecha_filtro={fecha_filtro}, contratista_filtro={contratista_filtro}")
    print(f"DEBUG: Página: {page}, Por página: {per_page}, Offset: {offset}")
    print(f"DEBUG: Query: {query}")
    print(f"DEBUG: Params: {params}")
    
    reportes = conn.execute(query, params).fetchall()
    print(f"DEBUG: Se encontraron {len(reportes)} reportes en esta página")
    
    # Obtener contratistas para filtro
    contratistas = conn.execute('''
        SELECT DISTINCT contratista 
        FROM rutas
        ORDER BY contratista
    ''').fetchall()
    
    conn.close()
    
    # Calcular información de paginación
    total_pages = (total_reportes + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    prev_page = page - 1 if has_prev else None
    next_page = page + 1 if has_next else None
    
    # Calcular rango de páginas para mostrar
    start_page = max(1, page - 2)
    end_page = min(total_pages, page + 2)
    
    pagination_info = {
        'page': page,
        'per_page': per_page,
        'total': total_reportes,
        'total_pages': total_pages,
        'has_prev': has_prev,
        'has_next': has_next,
        'prev_page': prev_page,
        'next_page': next_page,
        'start_page': start_page,
        'end_page': end_page,
        'pages': list(range(start_page, end_page + 1))
    }
    
    # Log de acceso al admin
    log_activity(current_user.id, 'access_admin', details=f'Acceso al panel de administración de rutas - Página {page}')
    
    return render_template('admin.html', 
                         reportes=reportes, 
                         contratistas=contratistas,
                         fecha_filtro=fecha_filtro,
                         contratista_filtro=contratista_filtro,
                         pagination=pagination_info)

@app.route('/update_reporte_status', methods=['POST'])
@login_required
def update_reporte_status():
    """Actualizar el estado de un reporte"""
    try:
        data = request.get_json()
        reporte_id = data['reporte_id']
        new_status = data['status']
        
        conn = get_db_connection()
        
        # Obtener reporte actual
        old_reporte = conn.execute(
            'SELECT estado, ruta_codigo, contratista FROM reportes_rutas WHERE id = ?',
            (reporte_id,)
        ).fetchone()
        
        if old_reporte:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE reportes_rutas 
                SET estado = ?
                WHERE id = ?
            ''', (new_status, reporte_id))
            
            conn.commit()
            
            # Log de actividad
            log_activity(current_user.id, 'update_reporte_status', 'reporte', reporte_id,
                        f'Ruta {old_reporte["ruta_codigo"]} ({old_reporte["contratista"]}): {old_reporte["estado"]} → {new_status}')
            
            conn.close()
            return jsonify({'success': True, 'message': 'Estado actualizado'})
        else:
            conn.close()
            return jsonify({'success': False, 'error': 'Reporte no encontrado'}), 404
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export_reportes')
@login_required
def export_reportes():
    """Exportar reportes a Excel"""
    from flask import make_response
    from io import BytesIO
    from datetime import datetime
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        fecha_filtro = request.args.get('fecha', datetime.now().strftime('%Y-%m-%d'))
        
        conn = get_db_connection()
        reportes = conn.execute('''
            SELECT r.*, ru.supervisor, ru.placa, ru.tipo
            FROM reportes_rutas r
            LEFT JOIN rutas ru ON r.ruta_id = ru.id
            WHERE DATE(r.fecha_reporte) = ?
            ORDER BY r.contratista, r.hora_aproximada_ingreso
        ''', (fecha_filtro,)).fetchall()
        conn.close()
        
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reportes Rutas {fecha_filtro}"
        
        # Encabezados
        headers = [
            'ID', 'Fecha', 'Hora Reporte', 'Contratista', 'Ruta', 'Supervisor',
            'Placa', 'Clientes Pendientes', 'Cajas en Camión', 'Hora Aprox. Ingreso',
            'Hora Exacta Envío', 'Ubicación Exacta', 'Comentarios', 'Estado', 'Reportado Por'
        ]
        
        # Agregar encabezados con estilo
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Agregar datos
        for row_idx, reporte in enumerate(reportes, 2):
            ws.cell(row=row_idx, column=1, value=reporte['id'])
            ws.cell(row=row_idx, column=2, value=reporte['fecha_reporte'])
            ws.cell(row=row_idx, column=3, value=reporte['hora_reporte'])
            ws.cell(row=row_idx, column=4, value=reporte['contratista'])
            ws.cell(row=row_idx, column=5, value=reporte['ruta_codigo'])
            ws.cell(row=row_idx, column=6, value=reporte['supervisor'] or '')
            ws.cell(row=row_idx, column=7, value=reporte['placa'] or '')
            ws.cell(row=row_idx, column=8, value=reporte['clientes_pendientes'])
            ws.cell(row=row_idx, column=9, value=reporte['cajas_camion'])
            ws.cell(row=row_idx, column=10, value=reporte['hora_aproximada_ingreso'])
            ws.cell(row=row_idx, column=11, value=reporte['hora_exacta_envio'] or '')
            ws.cell(row=row_idx, column=12, value=reporte['ubicacion_exacta'] or '')
            ws.cell(row=row_idx, column=13, value=reporte['comentarios'] or '')
            ws.cell(row=row_idx, column=14, value=reporte['estado'])
            ws.cell(row=row_idx, column=15, value=reporte['reportado_por'] or '')
        
        # Autoajustar columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Log de actividad
        log_activity(current_user.id, 'export_reportes', details=f'Exportó {len(reportes)} reportes del {fecha_filtro}')
        
        # Crear respuesta
        response = make_response(excel_buffer.getvalue())
        filename = f"reportes_rutas_{fecha_filtro.replace('-', '')}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
        
    except Exception as e:
        flash(f'Error exportando: {str(e)}', 'error')
        return redirect(url_for('admin'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login para administradores"""
    if current_user.is_authenticated:
        return redirect(url_for('admin'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Buscar usuario en la base de datos
        conn = get_db_connection()
        user_data = conn.execute(
            'SELECT * FROM users WHERE username = ? AND is_active = 1', 
            (username,)
        ).fetchone()
        conn.close()
        
        if user_data and check_password_hash(user_data['password_hash'], password):
            # Crear objeto User
            user = User(
                id=user_data['id'],
                username=user_data['username'],
                email=user_data['email'],
                role=user_data['role'],
                active=user_data['is_active']
            )
            
            # Iniciar sesión
            login_user(user)
            
            # Actualizar último login
            conn = get_db_connection()
            conn.execute(
                'UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = ?',
                (user_data['id'],)
            )
            conn.commit()
            conn.close()
            
            # Log de actividad
            log_activity(user_data['id'], 'login', details=f'Usuario {username} inició sesión')
            
            flash(f'¡Bienvenido {user_data["username"]}!', 'success')
            
            # Redirigir a la página solicitada o admin
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('admin'))
        else:
            # Log de intento fallido
            log_activity(None, 'failed_login', details=f'Intento de login fallido para usuario: {username}')
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Cerrar sesión"""
    if current_user.is_authenticated:
        # Log de actividad
        log_activity(current_user.id, 'logout', details=f'Usuario {current_user.username} cerró sesión')
        logout_user()
        flash('Sesión cerrada correctamente', 'info')
    
    return redirect(url_for('login'))

def log_activity(user_id, action, target_type=None, target_id=None, details=None):
    """Registrar actividad del usuario"""
    try:
        conn = get_db_connection()
        # Obtener IP del cliente
        ip_address = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
        
        conn.execute('''
            INSERT INTO activity_log (user_id, action, target_type, target_id, details, ip_address)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (user_id, action, target_type, target_id, details, ip_address))
        
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error logging activity: {e}")


def require_role(role):
    """Decorador para requerir rol específico"""
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            if role == 'admin' and not current_user.is_admin():
                flash('No tienes permisos para acceder a esta página', 'error')
                return redirect(url_for('admin'))
            elif role == 'supervisor' and not current_user.is_supervisor():
                flash('No tienes permisos para acceder a esta página', 'error')
                return redirect(url_for('admin'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/reload_rutas')
@login_required
@require_role('admin')
def reload_rutas():
    """Recargar rutas desde el archivo Excel"""
    if load_rutas_from_excel():
        flash('Rutas recargadas exitosamente desde Excel', 'success')
        log_activity(current_user.id, 'reload_rutas', details='Rutas recargadas desde Excel')
    else:
        flash('Error recargando rutas desde Excel', 'error')
    
    return redirect(url_for('admin'))

@app.route('/api/reportes')
def api_reportes():
    """API para obtener reportes"""
    fecha = request.args.get('fecha', datetime.now().strftime('%Y-%m-%d'))
    contratista = request.args.get('contratista', '')
    
    conn = get_db_connection()
    
    query = '''
        SELECT r.*, ru.supervisor, ru.placa, ru.tipo
        FROM reportes_rutas r
        LEFT JOIN rutas ru ON r.ruta_id = ru.id
        WHERE DATE(r.fecha_reporte) = ?
    '''
    params = [fecha]
    
    if contratista:
        query += ' AND r.contratista = ?'
        params.append(contratista)
    
    query += ' ORDER BY r.hora_reporte DESC'
    
    reportes = conn.execute(query, params).fetchall()
    conn.close()
    
    data = [dict(reporte) for reporte in reportes]
    return jsonify(data)

@app.route('/crear_reporte_prueba')
@login_required
def crear_reporte_prueba():
    """Crear un reporte de prueba para verificar el panel de administración"""
    try:
        conn = get_db_connection()
        
        # Obtener una ruta existente
        ruta = conn.execute('SELECT id, ruta, contratista FROM rutas LIMIT 1').fetchone()
        
        if not ruta:
            conn.close()
            flash('No hay rutas disponibles para crear un reporte de prueba.', 'error')
            return redirect(url_for('admin'))
        
        # Insertar reporte de prueba
        cursor = conn.cursor()
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        hora_actual = datetime.now().strftime('%H:%M')
        
        cursor.execute('''
            INSERT INTO reportes_rutas 
            (contratista, ruta_id, ruta_codigo, clientes_pendientes, 
             cajas_camion, hora_aproximada_ingreso, ubicacion_exacta, 
             latitud, longitud, hora_exacta_envio, comentarios, reportado_por,
             fecha_reporte)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            ruta['contratista'],
            ruta['id'],
            ruta['ruta'],
            5,  # clientes_pendientes
            10, # cajas_camion
            hora_actual,
            'Ubicación de prueba',
            10.123456,
            -84.123456,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Reporte de prueba creado automáticamente',
            current_user.username,
            fecha_actual
        ))
        
        conn.commit()
        reporte_id = cursor.lastrowid
        conn.close()
        
        flash(f'Reporte de prueba #{reporte_id} creado exitosamente para la fecha {fecha_actual}', 'success')
        return redirect(url_for('admin'))
        
    except Exception as e:
        flash(f'Error al crear reporte de prueba: {str(e)}', 'error')
        return redirect(url_for('admin'))

if __name__ == '__main__':
    print("🚀 Iniciando Sistema de Gestión de Rutas...")
    print(f"📂 Base de datos: {DATABASE}")
    
    # Inicializar base de datos
    init_db()
    
    # Cargar rutas desde Excel si existe el archivo
    if os.path.exists('DB_Rutas.xlsx'):
        print("📋 Cargando rutas desde Excel...")
        load_rutas_from_excel()
    else:
        print("⚠️ Archivo DB_Rutas.xlsx no encontrado")
    
    print("🌐 Aplicación lista")
    
    # Usar configuración dinámica de config.py
    from config import get_config
    
    config_class = get_config()
    config_instance = config_class()
    
    print(f"🌐 Servidor iniciando en puerto: {config_instance.PORT}")
    print(f"🔧 Debug mode: {config_instance.DEBUG}")
    print(f"🏠 Host: {config_instance.HOST}")
    print(f"🌍 PORT env var: '{os.environ.get('PORT', 'NOT_SET')}'")
    print(f"🚂 Railway env: '{os.environ.get('RAILWAY_ENVIRONMENT', 'NOT_SET')}'")
    
    if config_instance.DEBUG:
        print(f"   Local: http://127.0.0.1:{config_instance.PORT}")
        print("👤 Admin: admin / admin123")
    else:
        print("🚂 Modo producción activado para Railway")
    
    # Ejecutar la aplicación
    app.run(
        debug=config_instance.DEBUG, 
        host=config_instance.HOST, 
        port=config_instance.PORT
    )
